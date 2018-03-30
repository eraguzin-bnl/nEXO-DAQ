from scripts.sbnd_femb_meas import FEMB_DAQ
import os
import sys
#import copy
from datetime import datetime
import pickle
import logging
import time
import shutil
import glob
import queue
import threading
from user_settings import user_editable_settings
import numpy as np
import openpyxl as px
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Font, Side, PatternFill
from openpyxl.utils import get_column_letter
settings = user_editable_settings()

class main:
    def loop(self):
        #Starts by writing a test value to both FPGAs to make sure they respond
        print ("Start")
        sys.stdout.flush()
        for i in range(settings.chip_num):
            self.sbnd.femb_config.femb.init_ports(hostIP = settings.CHIP_IP[i],
                                                  destIP = self.sbnd.femb_config.femb.UDP_IP[i], dummy_port = i)
        self.sbnd.femb_config.femb.write_reg(8, 6, board = "wib")
        response = self.sbnd.femb_config.femb.read_reg(8, board = "wib")
        if (response != 6):
            print ("There is no communication with the WIB.  Try resetting the power, making sure all connections "
                   "are correct, and restarting the program.")
            print (response)
            
        self.sbnd.femb_config.femb.write_reg(8, 7, board = "femb")
        time.sleep(0.1)
        response = self.sbnd.femb_config.femb.read_reg(8, board = "femb")
        if (response != 7):
            print ("There is no communication with the FPGA.  Try resetting the power, making sure all connections "
                   "are correct, and restarting the program.")
            print (response)
            
        else:
            self.sbnd.femb_config.resetFEMBBoard()
            self.sbnd.femb_config.resetWIBBoard()
            self.sbnd.femb_config.initBoard()
            self.sbnd.femb_config.syncADC([0,1,2,3])
            print ("Communication is working with the FPGA.")

            
        self.sbnd.femb_config.fe_reg.info.fe_regs_sw = self.sbnd.femb_config.fe_reg.REGS
        self.sbnd.femb_config.adc_reg.info.adc_regs_sw = self.sbnd.femb_config.adc_reg.REGS
        print ("\n\n-----------------------------------------------------------------------------")
        print ("BNL Cold ASIC Python Program for nEXO")
        print ("Uses Version 7 of the Front End and ADC chips")
        
        print ("The directory that data is saved to will be")
        print (settings.path)
        self.chip = 1
        self.chn = 0
        try: 
            os.makedirs(settings.path)
        except OSError:
            if os.path.exists(settings.path):
                pass
#        self.plot_pulsed_data_no_change()
        while(1):
            #self.help_info()
            raw_in = input("Enter your input\n")
            if (raw_in == "end"):
                break
            elif (raw_in == "live"):
                for i in range(settings.chip_num):
                    print ("Live packet of Chip {}...".format(i))
                    self.plot_live_packet(chip = i)
                
            elif (raw_in == "pulsed"):
                print ("Pulsed packet...")
                self.plot_pulsed_packets(source = "int")
                
            elif (raw_in == "pulsed2"):
                print ("Pulsed packet...")
                self.plot_pulsed_data_no_change()
                
            elif (raw_in == "setup"):
                print ("Setting up...")
                self.setup_no_change()
                
            elif (raw_in == "sync"):
                print ("Syncing...")
                self.sbnd.femb_config.syncADC([0,1,2,3])
                
            elif (raw_in == "temp"):
                print ("Polling for the temperature")
                self.sbnd.femb_config.syncADC([0,1,2,3])
                self.check_temp([0,1,2,3])

            elif (raw_in == "bathtub"):
                print ("Bathtub Curve")
                filename = settings.path + input("Enter spreadsheet name\n") + ".xlsx"
                self.bathtub([0,1,2,3], filename)

            elif (raw_in == "noise"):
                print ("Noise Calibration...")
                self.initialize_calibration("Noise")
                self.noise_calibration()
                self.noise = True
                self.save_config()

            elif (raw_in == "fpga"):
                print ("Gain calibration (with external FPGA pulse)...")
                self.initialize_calibration("FPGA gain calibration")
                self.gain_calibration_fpga()
                self.fpga = True
                self.save_config()

            elif (raw_in == "int"):
                print ("Gain calibration (with internal DAC pulse)...")
                self.initialize_calibration("Internal gain calibration")
                self.gain_calibration_internal()
                self.int = True
                self.save_config()

            elif (raw_in == "full"):
                print ("Full calibration...")
                self.initialize_calibration("Full calibration")
                self.noise_calibration()
                self.noise = True
                self.save_config()
#                self.gain_calibration_fpga()
#                self.fpga = True
#                self.save_config()
                self.gain_calibration_internal()
                self.int = True
                self.save_config()
                print ("Full calibration complete")
                
            elif (raw_in == "trigger"):
                self.triggered_data()
                print ("Finished collecting triggered data")
                
            elif (raw_in == "help"):
                self.help_info()
                
            else:
                print ("That's not a function")

        
        #Noise measurement
    def noise_calibration(self):
        print ("Noise measurement beginning...")
        for snc in self.snc_tuple:
            for sg in self.sg_tuple:
                for st in self.st_tuple:
                    self.sbnd.femb_config.fe_reg.set_fe_board() # reset the registers value
                    self.sbnd.femb_config.fe_reg.set_fe_board(sts=1, snc=snc[0], sg=sg[0], st=st[0],
                                                              sdacsw1=0, sdacsw2=0, sdac = 45)
                    PattenDir = sg[1] + "_" + st[1] + "_" + snc[1]
                    print (PattenDir)
    
                    self.sbnd.testpattern = PattenDir
                    self.sbnd.make_dir(self.cali_path)
                    
                    self.sbnd.femb_config.configFeAsic(output = "suppress")
                    print ("Optimizing ADC offset values")
                    self.sbnd.femb_config.optimize_offset(self.sbnd.path,sg[0])
    
                    self.sbnd.save_rms_noise()
                    
                    data = {'base': snc[1],
                            'gain': sg[1],
                            'peak': st[1],
                            'temp': settings.temp,
                            }
                    with open(self.sbnd.path + 'chip_settings.cfg', 'wb') as f:
                        pickle.dump(data, f, pickle.HIGHEST_PROTOCOL)
                    
        print ("Noise measurement finished!")
        
    def gain_calibration_fpga(self):
        print ("Gain measurement(FPGA) beginning...")
        for sg in self.sg_tuple:
            for st in self.st_tuple:
                self.sbnd.femb_config.fe_reg.set_fe_board() # reset the registers value
                self.sbnd.femb_config.fe_reg.set_fe_board(sts=1,sg=sg[0], st=st[0], snc=1,
                                                          sdacsw1=1, sdacsw2=0)
                PattenDir = sg[1] + "_" + st[1] + "_" + self.snc_tuple[1][1]
                print (PattenDir)
    
                self.sbnd.testpattern = PattenDir
                self.sbnd.make_dir(self.cali_path)
                self.sbnd.fpga_dac_cali()
                
                data = {'base': self.snc_tuple[1][1],
                            'gain': sg[1],
                            'peak': st[1],
                            'temp': settings.temp,
                            }
                with open(self.sbnd.path + 'chip_settings.cfg', 'wb') as f:
                    pickle.dump(data, f, pickle.HIGHEST_PROTOCOL)
                        
        print ("Gain measurement(FPGA) finished!")

    def gain_calibration_internal(self):
        print ("Gain measurement(Internal) beginning...")
#        ok = False
#        while (ok == False):
#            self.plot_pulsed_data_no_change()
#            answer = input("Do the pulses look ok? y/n\n")
#            if (answer == "y"):
#                ok = True
#            else:
#                self.sbnd.femb_config.configAdcAsic()
        for sg in self.sg_tuple:
            for st in self.st_tuple:
                self.sbnd.current_sg = sg
                self.sbnd.current_st = st
                self.sbnd.femb_config.fe_reg.set_fe_board() # reset the registers value
                self.sbnd.femb_config.fe_reg.set_fe_board(sts=1,sg=sg[0], st=st[0], snc=1,
                                                          sdacsw1=0, sdacsw2=1, sdac=5)
                PattenDir = sg[1] + "_" + st[1] + "_" + self.snc_tuple[1][1]
                print (PattenDir )
            
                self.sbnd.testpattern = PattenDir
                self.sbnd.make_dir(self.cali_path)
                self.sbnd.intdac_cali()
                
                data = {'base': self.snc_tuple[1][1],
                            'gain': sg[1],
                            'peak': st[1],
                            'temp': settings.temp,
                            }
                with open(self.sbnd.path + 'chip_settings.cfg', 'wb') as f:
                    pickle.dump(data, f, pickle.HIGHEST_PROTOCOL)
                        
        print ("Gain measurement(Internal) finished!")
        
    def plot_live_packet(self, chip, packets = 25):
        
        data = self.sbnd.femb_config.femb.get_data_packets(ip = settings.CHIP_IP[chip], 
                                                           data_type = "int", num = packets, header = False)
        self.sbnd.analyze.UnpackData(path = "data", data = data)
        
    def plot_pulsed_data_no_change(self):
        for chip in range(settings.chip_num):
            reg_5_original = self.sbnd.femb_config.femb.read_reg(self.sbnd.femb_config.REG_TEST_PULSE, "femb")
            reg_5_value = ((150<<16)&0xFFFF0000) + ((reg_5_original)& 0xFFFF)   
            self.sbnd.femb_config.femb.write_reg(self.sbnd.femb_config.REG_TEST_PULSE, reg_5_value, "femb")
            
            #Enable test pulses
            self.sbnd.femb_config.femb.write_reg(18, 0x0, board = "femb")
            
            #Enable internal test pulse
            self.sbnd.femb_config.femb.write_reg(16, 0x002, board = "femb")
            
            time.sleep(0.01)
    
            data = self.sbnd.femb_config.femb.get_data_packets(ip = settings.CHIP_IP[chip], 
                                                               data_type = "int", num = 25, header = False)
            print ("Chip {}".format(chip))
            self.sbnd.analyze.UnpackData(path = "data", data = data)
                
            #Bring everything back the way it was            
            self.sbnd.femb_config.femb.write_reg(self.sbnd.femb_config.REG_TEST_PULSE, reg_5_original, "femb")
            
            #Disable all test pulses
            self.sbnd.femb_config.femb.write_reg(18, 0x1, board = "femb")
            
            #Disable FPGA test pulse
            self.sbnd.femb_config.femb.write_reg(16, 0x1, board = "femb")

    def bathtub(self, adc, filename):
        wb = Workbook()
        
#        redFill = PatternFill(start_color='FFC7CE', end_color='CEC7FF', fill_type='solid')
        Fill5 = PatternFill(start_color='05FF30', end_color='05FF30', fill_type='solid')
        Fill4 = PatternFill(start_color='B1FF08', end_color='05FF30', fill_type='solid')
        Fill3 = PatternFill(start_color='F5D708', end_color='05FF30', fill_type='solid')
        Fill2 = PatternFill(start_color='F59008', end_color='05FF30', fill_type='solid')
        Fill1 = PatternFill(start_color='FA4D07', end_color='05FF30', fill_type='solid')
        Fill0 = PatternFill(start_color='FF0505', end_color='05FF30', fill_type='solid')
#        redFont = Font(color='9C0006')

        first = True
        
        self.sbnd.femb_config.adc_reg.set_adc_board(d=0, pcsr=1, pdsr=1, slp=0, tstin=0,
                                                clk = 0, frqc = 0, en_gr = 1, f0 = 0, f1 = 0, 
                                                f2 = 0, f3 = 0, f4 = 1, f5 = 0, slsb = 0, show="FALSE")        
        
        reg3 = self.sbnd.femb_config.femb.read_reg(3, board = "femb")
        newReg3 = ( reg3 | 0x80000000 )
        self.sbnd.femb_config.femb.write_reg( 3, newReg3, board = "femb") #31 - enable ADC test pattern
        
        time.sleep(0.01)

        initLATCH1_4 = self.sbnd.femb_config.femb.read_reg( self.sbnd.femb_config.REG_LATCHLOC1_4, board = "femb" )
        
        for adcNum in adc:
            print ("-------------------------------------------------")
            print ("Bathtub Curve Test --> Testing ADC {}".format(adcNum))

            if (adcNum == 0):
                initPLL_read = self.sbnd.femb_config.femb.read_reg( 21, board = "femb" )
                initPLL_clock = self.sbnd.femb_config.femb.read_reg( 24, board = "femb" )
            elif (adcNum == 1):
                initPLL_read = self.sbnd.femb_config.femb.read_reg( 21, board = "femb" )
                initPLL_clock = self.sbnd.femb_config.femb.read_reg( 24, board = "femb" )
            elif (adcNum == 2):
                initPLL_read = self.sbnd.femb_config.femb.read_reg( 22, board = "femb" )
                initPLL_clock = self.sbnd.femb_config.femb.read_reg( 25, board = "femb" )
            elif (adcNum == 3):
                initPLL_read = self.sbnd.femb_config.femb.read_reg( 22, board = "femb" )
                initPLL_clock = self.sbnd.femb_config.femb.read_reg( 25, board = "femb" )
                
            initPLL_read2 = self.sbnd.femb_config.femb.read_reg( 23, board = "femb" )
            initPLL_clock2 = self.sbnd.femb_config.femb.read_reg( 26, board = "femb" )
                          
            print ("Bathtub Curve Test --> Initial Read PLL is {}".format(hex(initPLL_read)))
            
            initSetting = (initLATCH1_4 & (0xFF << (8 * adcNum))) >> (8 * adcNum)
            
            print ("Bathtub Curve Test --> Initial Latch is {}".format(hex(initLATCH1_4)))
    
            print ("Bathtub Curve Test --> First testing around the initial setting of {}".format(hex(initSetting)))
            
            print ("Bathtub Curve Test --> initPLL_read2 is {}".format(hex(initPLL_read2)))
            
            print ("Bathtub Curve Test --> initPLL_clock2 is {}".format(hex(initPLL_clock2)))
            
            print ("Bathtub Curve Test --> Will test ({},{},{}) for the Read clock and ({},{},{}) for the Write clock".format(
                    settings.bath_read_min[adcNum], settings.bath_read_max[adcNum], settings.bath_read_step[adcNum],
                    settings.bath_write_min[adcNum], settings.bath_write_max[adcNum], settings.bath_write_step[adcNum]))
            
            for shift in range(initSetting - settings.bath_latch_down[adcNum], initSetting + 1 + settings.bath_latch_up[adcNum], 1):
                shiftMask = (0x3F << 8*adcNum)
                testShift = ( (initLATCH1_4 & ~(shiftMask)) | (shift << 8*adcNum) )
                self.sbnd.femb_config.femb.write_reg( self.sbnd.femb_config.REG_LATCHLOC1_4, testShift, board = "femb" )
                
                if (first == True):
                    ws = wb.active
                    ws.title = "ASIC {}, Latch {}".format(adcNum, hex(shift))
                    first = False
                else:
                    ws = wb.create_sheet("ASIC {}, Latch {}".format(adcNum, hex(shift)))
                
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=34)
                ws.cell(row=1, column=1).value = "ASIC {}, Latch {}".format(adcNum, hex(shift))
                ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
                ws.cell(row=2, column=1).value = "Read Clock"
                ws.merge_cells(start_row=36, start_column=2, end_row=36, end_column=34)
                ws.cell(row=36, column=2).value = "Write Clock"
                ws.cell(row=36, column=2).alignment = Alignment(horizontal='center')
                for i in range(32):
                    ws.cell(row = i + 3, column = 1).value = hex(i)
                    ws.cell(row = 35, column = i + 2).value = hex(i)
                    ws.column_dimensions[get_column_letter(i+2)].width = 5
        
                wb.save(filename = filename)
                for PLL_read in range(settings.bath_read_min[adcNum], settings.bath_read_max[adcNum], settings.bath_read_step[adcNum]):
                    absolute = abs(PLL_read)
                    
                    if (adcNum == 0):
                        self.sbnd.femb_config.femb.write_reg( 21, (initPLL_read & 0xFFFF0000) + absolute, board = "femb" )
                        if (PLL_read > 0):
                            pll2 = initPLL_read2 & 0x000E0000
                        else:
                            pll2 = (initPLL_read2 | 0x00010000) & 0x7FFFFFF
    #                    print ("Init PLL is {}, adjusted PLL is {} and adding {} to it gives {}".format(hex(initPLL),
    #                       hex(initPLL & 0xFFFF0000), hex(PLL), hex((initPLL & 0xFFFF0000) + PLL)))
                    elif (adcNum == 1):
                        self.sbnd.femb_config.femb.write_reg( 21, (initPLL_read & 0xFFFF) + (absolute << 16), board = "femb" )
                        if (PLL_read > 0):
                            pll2 = initPLL_read2 & 0x000D0000
                        else:
                            pll2 = (initPLL_read2 | 0x00020000) & 0x7FFFFFF
    #                    print ("Init PLL is {}, adjusted PLL is {} and adding {} to it gives {}".format(hex(initPLL),
    #                       hex(initPLL & 0xFFFF), hex(PLL << 16), hex((initPLL & 0xFFFF) + (PLL << 16))))
                    elif (adcNum == 2):
                        self.sbnd.femb_config.femb.write_reg( 22, (initPLL_read & 0xFFFF0000) + absolute, board = "femb" )
                        if (PLL_read > 0):
                            pll2 = initPLL_read2 & 0x000B0000
                        else:
                            pll2 = (initPLL_read2 | 0x00040000) & 0x7FFFFFF
    #                    print ("Init PLL is {}, adjusted PLL is {} and adding {} to it gives {}".format(hex(initPLL),
    #                       hex(initPLL & 0xFFFF0000), hex(PLL), hex((initPLL & 0xFFFF0000) + PLL)))
                    elif (adcNum == 3):
                        self.sbnd.femb_config.femb.write_reg( 22, (initPLL_read & 0xFFFF) + (absolute << 16), board = "femb" )
                        if (PLL_read > 0):
                            pll2 = initPLL_read2 & 0x00070000
                        else:
                            pll2 = (initPLL_read2 | 0x00080000) & 0x7FFFFFF
    #                    print ("Init PLL is {}, adjusted PLL is {} and adding {} to it gives {}".format(hex(initPLL),
    #                       hex(initPLL & 0xFFFF), hex(PLL << 16), hex((initPLL & 0xFFFF) + (PLL << 16))))
    
                    
                    self.sbnd.femb_config.femb.write_reg ( 23, pll2, "femb")
                    self.sbnd.femb_config.femb.write_reg ( 23, 0x80000000 + pll2, "femb")
                    
                    for PLL_clock in range(settings.bath_write_min[adcNum], settings.bath_write_max[adcNum], settings.bath_write_step[adcNum]):
                        absolute = abs(PLL_clock)
                        
                        if (adcNum == 0):
                            self.sbnd.femb_config.femb.write_reg( 24, (initPLL_clock & 0xFFFF0000) + absolute, board = "femb" )
                            if (PLL_clock > 0):
                                pll3 = initPLL_clock2 & 0x000E0000
                            else:
                                pll3 = (initPLL_clock2 | 0x00010000) & 0x7FFFFFF
                                       
        #                    print ("Init PLL is {}, adjusted PLL is {} and adding {} to it gives {}".format(hex(initPLL),
        #                       hex(initPLL & 0xFFFF0000), hex(PLL), hex((initPLL & 0xFFFF0000) + PLL)))
                        elif (adcNum == 1):
                            self.sbnd.femb_config.femb.write_reg( 24, (initPLL_clock & 0xFFFF) + (absolute << 16), board = "femb" )
                            if (PLL_clock > 0):
                                pll3 = initPLL_clock2 & 0x000D0000
                            else:
                                pll3 = (initPLL_clock2 | 0x00020000) & 0x7FFFFFF
        #                    print ("Init PLL is {}, adjusted PLL is {} and adding {} to it gives {}".format(hex(initPLL),
        #                       hex(initPLL & 0xFFFF), hex(PLL << 16), hex((initPLL & 0xFFFF) + (PLL << 16))))
                        elif (adcNum == 2):
                            self.sbnd.femb_config.femb.write_reg( 25, (initPLL_clock & 0xFFFF0000) + absolute, board = "femb" )
                            if (PLL_clock > 0):
                                pll3 = initPLL_clock2 & 0x000B0000
                            else:
                                pll3 = (initPLL_clock2 | 0x00040000) & 0x7FFFFFF
        #                    print ("Init PLL is {}, adjusted PLL is {} and adding {} to it gives {}".format(hex(initPLL),
        #                       hex(initPLL & 0xFFFF0000), hex(PLL), hex((initPLL & 0xFFFF0000) + PLL)))
                        elif (adcNum == 3):
                            self.sbnd.femb_config.femb.write_reg( 25, (initPLL_clock & 0xFFFF) + (absolute << 16), board = "femb" )
                            if (PLL_clock > 0):
                                pll3 = initPLL_clock2 & 0x00070000
                            else:
                                pll3 = (initPLL_clock2 | 0x00080000) & 0x7FFFFFF
        #                    print ("Init PLL is {}, adjusted PLL is {} and adding {} to it gives {}".format(hex(initPLL),
        #                       hex(initPLL & 0xFFFF), hex(PLL << 16), hex((initPLL & 0xFFFF) + (PLL << 16))))
    
                        
                        self.sbnd.femb_config.femb.write_reg ( 26, pll3, "femb")
                        self.sbnd.femb_config.femb.write_reg ( 26, 0x80000000 + pll3, "femb")
                        
                        reg21 = self.sbnd.femb_config.femb.read_reg( 21, board = "femb" )
                        reg22 = self.sbnd.femb_config.femb.read_reg( 22, board = "femb" )
                        reg23 = self.sbnd.femb_config.femb.read_reg( 23, board = "femb" )
                        reg24 = self.sbnd.femb_config.femb.read_reg( 24, board = "femb" )
                        reg25 = self.sbnd.femb_config.femb.read_reg( 25, board = "femb" )
                        reg26 = self.sbnd.femb_config.femb.read_reg( 26, board = "femb" )
                        
                        print ("Bathtub Curve Test --> Testing with Reg 21 = {}, Reg 22 = {}, Reg 23 = {}".format(hex(reg21), hex(reg22), hex(reg23)))
                        print ("Bathtub Curve Test --> Testing with Reg 24 = {}, Reg 25 = {}, Reg 26 = {}".format(hex(reg24), hex(reg25), hex(reg26)))
                
                        
                        works = 0
                        for tries in range(5):
                            self.sbnd.femb_config.femb.write_reg ( self.sbnd.femb_config.REG_ASIC_RESET, 1, "femb")
                            self.sbnd.femb_config.configAdcAsic(output = "suppress")
                            time.sleep(0.05)
                            if (self.sbnd.femb_config.testUnsync(adcNum, output = "suppress") == 0):
                                works = works + 1
    
                        print ("Bathtub Curve Test --> Result is {}".format(works))
                        ws.cell(row = PLL_read + 3, column = PLL_clock + 2).value = "{}".format(works)
                        
                        if (works == 5):
                            ws.cell(row = PLL_read + 3, column = PLL_clock + 2).fill = Fill5
                        elif (works == 4):
                            ws.cell(row = PLL_read + 3, column = PLL_clock + 2).fill = Fill4
                        elif (works == 3):
                            ws.cell(row = PLL_read + 3, column = PLL_clock + 2).fill = Fill3
                        elif (works == 2):
                            ws.cell(row = PLL_read + 3, column = PLL_clock + 2).fill = Fill2
                        elif (works == 1):
                            ws.cell(row = PLL_read + 3, column = PLL_clock + 2).fill = Fill1
                        elif (works == 0):
                            ws.cell(row = PLL_read + 3, column = PLL_clock + 2).fill = Fill0
                        
                        wb.save(filename = filename)

        #Bring everything back the way it was            
        self.sbnd.femb_config.femb.write_reg( 3, (reg3&0x7fffffff), board = "femb" )
        
    def setup_no_change(self):
        
        self.sbnd.femb_config.adc_reg.set_adc_board(d=0, pcsr=1, pdsr=1, slp=0, tstin=0,
                 clk = 0, frqc = 0, en_gr = 1, f0 = 0, f1 = 0, 
                 f2 = 0, f3 = 0, f4 = 1, f5 = 0, slsb = 0, show="FALSE")        
        
        self.sbnd.femb_config.fe_reg.set_fe_board(sts=1,sg=3, sbf = 1, st=1, snc=0,
                                                          sdacsw1=0, sdacsw2=1, sdac=5)
        
        for i in range (settings.chip_num):
            self.sbnd.femb_config.fe_reg.set_fe_chn(chip = i, chn = 1,sts=1,sg=1, sbf = 1, st=2, snc=0)
        
#        self.sbnd.femb_config.configAdcAsic()
        self.sbnd.femb_config.configFeAsic()
        
    def check_temp(self,chips):
        self.sbnd.femb_config.fe_reg.set_fe_board(sts=1, snc=1, sg=2, st=1, smn=0, sbf=1, 
                           slk = 0, stb = 2, s16=0, slkh=0, sdc=0, sdacsw2=0, sdacsw1=0, sdac=0, show=False)
        
        self.sbnd.femb_config.adc_reg.set_adc_board(d=0, pcsr=0, pdsr=0, slp=0, tstin=0,
                 clk = 0, frqc = 0, en_gr = 0, f0 = 0, f1 = 0, 
                 f2 = 0, f3 = 0, f4 = 1, f5 = 0, slsb = 0, show="FALSE")
        
        self.sbnd.femb_config.configAdcAsic()
        self.sbnd.femb_config.configFeAsic()
        
        for i in chips:
            data = self.sbnd.femb_config.get_data_chipXchnX(chip = i, chn = 0, packets = 1)
            average_in_bits = np.mean(data)
            #12 bit ADC from 0.2 - 1.6 V.  1.6-0.2/4096 = 0.342 for slop and 200 is the offset
            average_in_mV = (0.342 * average_in_bits) + 200
            #Temperature output on average is 895 mV at 300 K and 250 mV at 77 K
            #That's about 0.3457 mV/K with an offset of -9.434
            average_in_K = (0.1221 * average_in_bits) - 15.97
            print("Debug info: Bits - {}, mV - {}".format(average_in_bits, average_in_mV))
            print("Temperature according to Chip {} is {}".format(i, average_in_K))

        
    def plot_pulsed_packets(self, source = "int", packets = 5):
        
        #Increase pulse frequency and the packet size so you have to see a pulse
        
        reg_5_original = self.sbnd.femb_config.femb.read_reg(self.sbnd.femb_config.REG_TEST_PULSE, "femb")
        reg_5_value = ((75<<16)&0xFFFF0000) + ((reg_5_original)& 0xFFFF)    
        self.sbnd.femb_config.femb.write_reg(self.sbnd.femb_config.REG_TEST_PULSE, reg_5_value, "femb")

        
        #Enable test pulses
        self.sbnd.femb_config.femb.write_reg(18, 0x0, board = "femb")
        
        self.sbnd.femb_config.adc_reg.set_adc_board(d=0, pcsr=1, pdsr=1, slp=0, tstin=0,
                 clk = 0, frqc = 0, en_gr = 0, f0 = 0, f1 = 0, 
                 f2 = 0, f3 = 0, f4 = 1, f5 = 0, slsb = 0, show="FALSE")
        
#        self.sbnd.femb_config.configAdcAsic()
            
        if (source == "int"):
        
            self.sbnd.femb_config.fe_reg.set_fe_board(sts=1, snc=0, sg=2, st=1, smn=0, sbf=1, 
                           slk = 0, stb = 0, s16=0, slkh=0, sdc=0, sdacsw2=1, sdacsw1=0, sdac=15, show=False)
            
#            self.sbnd.femb_config.fe_reg.set_fe_chn(chip=1, 
#                    chn=0, sts=1, snc=0, sg=2, st=1, smn=1, sbf=0, show = "FALSE")
                
            self.sbnd.femb_config.configFeAsic()
            
            #Enable internal test pulse
            self.sbnd.femb_config.femb.write_reg(16, 0x002, board = "femb")
            
        elif (source == "fpga"):
            
            #Enable FPGA test pulse
            self.sbnd.femb_config.femb.write_reg(16, 0x101, board = "femb")
            
        else:
            
            print ("PULSE_PLOT --> Pulse source must be 'int' or 'fpga'")
            

        #Enable External test pulse
        #udp.write_reg(16, 0x0, "femb")

        time.sleep(0.01)
        for i in range(settings.chip_num):
            data = self.sbnd.femb_config.femb.get_data_packets(ip = settings.CHIP_IP[i], 
                                                               data_type = "int", num = packets, header = False)
            print ("Chip {}".format(i))
            self.sbnd.analyze.UnpackData(path = "data", data = data)
            
        #Bring everything back the way it was            
        self.sbnd.femb_config.femb.write_reg(self.sbnd.femb_config.REG_TEST_PULSE, reg_5_original, "femb")
        
        #Disable all test pulses
        self.sbnd.femb_config.femb.write_reg(18, 0x1, board = "femb")
        
        #Disable FPGA test pulse
        self.sbnd.femb_config.femb.write_reg(16, 0x1, board = "femb")
        
    def triggered_data(self):
        
        note = input( "Enter a test note (e.g. 'Test Run' or '004'):\n")
        self.triggered_path = settings.path + "Triggered_{}\\".format(note)
        
        frame_size = 0x0f00
        if (frame_size%13 != 0):
            frame_size = 13 * (frame_size//13)
            
        trigger_mode = 1;
        reg31_value = ((trigger_mode<<31) & 0x80000000) + frame_size
        self.sbnd.femb_config.femb.write_reg(31, reg31_value, "wib")
        
        self.sbnd.femb_config.femb.write_reg(22, settings.timeout_wait, "wib")

        total_length = settings.post_trigger + settings.pre_trigger
        
        reg21_value = (((settings.pre_trigger<<16) & 0xFF0000) +
                      ((total_length<<8) & 0xFF00) +
                      ((settings.default_delay) & 0xFF ))
                                 
        self.sbnd.femb_config.femb.write_reg(21, reg21_value, "wib")
        
        
        
        self.sbnd.femb_config.fe_reg.set_fe_board(sts=1, snc=0, sg=3, st=1, smn=0, sbf=1, 
                       slk = 0, stb = 0, s16=0, slkh=0, sdc=0, sdacsw2=0, sdacsw1=1, sdac=0, show="True")
        
        self.sbnd.femb_config.adc_reg.set_adc_board(d=0, pcsr=1, pdsr=1, slp=0, tstin=0,
                 clk = 0, frqc = 0, en_gr = 0, f0 = 0, f1 = 0, 
                 f2 = 0, f3 = 0, f4 = 0, f5 = 0, slsb = 0, show="True")
        
        self.sbnd.femb_config.configAdcAsic()
        
        self.sbnd.femb_config.configFeAsic()
        time.sleep(0.1)
        
        #Enable test pulses
        self.sbnd.femb_config.femb.write_reg(18, 0x0, board = "femb")
        
        #Enable external test pulse
        self.sbnd.femb_config.femb.write_reg(16, 0x0, board = "femb")
        
        logging.basicConfig(level=logging.DEBUG,
                    format='[%(levelname)s] (%(threadName)-10s) %(message)s',
                    )
        if (os.path.isdir(self.triggered_path)):
            print ("Deleting old files...")
            for the_file in os.listdir(self.triggered_path):
                file_path = os.path.join(self.triggered_path, the_file)
                try:
                    if os.path.isfile(file_path):
                        os.unlink(file_path)
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    print(e)
        
        time.sleep(0.5)
        
        try: 
            os.makedirs(self.triggered_path)
        except OSError:
            if os.path.exists(self.triggered_path):
                pass
            
        print ("Packets will be saved in {}".format(self.triggered_path))
        
        file_rec = self.triggered_path + "Registers_During_Test.txt"
        screendisplay = sys.stdout
        sys.stdout = open(file_rec, "a+")
        
        for i in range(settings.chip_num):
            self.sbnd.femb_config.fe_reg.info.fe_chip_status(i, status = "sent")
            self.sbnd.femb_config.adc_reg.info.adc_chip_status(i, status = "sent")
            
        sys.stdout.close()
        sys.stdout = screendisplay

        threads = []

        q = queue.Queue(1)
        
        num_of_packets = int(input("Enter number of packets to collect (if trigger is running at 1 kHz, "
                               "6000 packets will be 1 minute, etc...).  '0' means go for a specific time"
                                " or until the user decides to end manually\n"))
        
        collection_seconds = int(input("Enter time limit in seconds.  The data collection will stop once"
                                   " the limit is hit\n"))
        
        check_increment = float(input("Enter how often you want updates in seconds (or 0 if you want no updates).\n"))
        
        
        t1 = threading.Thread(name='C DLL Thread', 
                              target=self.listen_for_trigger_data, args=(num_of_packets, 
                                settings.chip_num, 
                                settings.packets_per_file,
                                settings.buffer_size,
                                self.sbnd.femb_config.femb.PORT_HSDATA,
                                settings.CHIP_IP[0],
                                settings.CHIP_IP[1],
                                settings.CHIP_IP[2],
                                settings.CHIP_IP[3],
                                self.triggered_path,
                                q)
                              )
        threads.append(t1)
        
        #Packet num reset
        self.sbnd.femb_config.femb.write_reg(0, int('100000',2), "wib")
        
        t1.start()
        
        dll_handle = q.get()
        
        #Thought this would allow q.join() to pass in thread, but it doesn't
        q.task_done()
        
        start_time_seconds = time.time()  # Get timezone naive now
        check_time = start_time_seconds
        
        while (t1.is_alive()):
            if (time.time() < (check_time + check_increment) or (check_increment == 0)):
                time.sleep(1)
            elif (time.time() > start_time_seconds + collection_seconds):
                end_function = dll_handle.end_data_collection
                result = end_function()
                if (result == 0):
                    print ("Successfully told DLL to quit")
                    break
            else:
                print ("Still going at {}".format(datetime.now().strftime('%H:%M:%S')))
                check_time = time.time()

        t1.join()
        newest = [[],[],[],[]]
        print ("Separating Packets...")
        separated_path = self.sbnd.analyze.Seperate_Packets(self.triggered_path)
        for i in range(settings.chip_num):
            string = separated_path + '*' + str(i) + '_Packet*.dat'
            newest[i] = max(glob.iglob(string), key=os.path.getctime)
        
            print("Chip " + str(i) + " plot displayed is")
            print (newest[i])
            
            self.sbnd.analyze.UnpackData(path = "default", data = newest[i], return_data=False)
        
    def help_info(self):
        print ("Type in the function you want to call.")
        print ("Type 'live' to see a plotted packet for each chip.")
        print ("Type 'pulsed' to see a plotted packet for each chip with an internal test pulse.")
        print ("Type 'noise' to begin noise calibration.")
        print ("Type 'fpga' to begin FPGA gain calibration.")
        print ("Type 'int' to begin internal gain calibration.")
        print ("Type 'full' to begin all three calibrations.")
        print ("Type 'trigger' to read triggered data.")
        print ("Type 'bathtub' to do an in depth cold sync test (For the long cable only)")
        print ("Type 'end' to exit.")
        
    def initialize_calibration(self, source = ""):
        print ("Initialize Calibration")
        
        note = input( "Enter a test note (e.g. 'Test Run' or '004'):\n")
        self.cali_path = settings.path + "Calibration_{}\\".format(note)
        
        try: 
            os.makedirs(self.cali_path)
        except OSError:
            if os.path.exists(self.cali_path):
                pass
            
        print ("Test will assume the temperature is {}".format(settings.temp))
        
        try:
            with open(self.cali_path + 'configuration.cfg', 'rb') as f:
                data = pickle.load(f)
                self.noise = data['noise']
                self.fpga = data['fpga']
                self.int = data['int']
                
        except FileNotFoundError:
            data = {'temp': settings.temp,
                    'noise': False,
                    'fpga': False,
                    'int': False
                    }
            
        with open(self.cali_path + 'configuration.cfg', 'wb') as f:
            pickle.dump(data, f, pickle.HIGHEST_PROTOCOL)
        
        
        print ("Beginning Calibration...")
        screendisplay = sys.stdout
        sys.stdout = open(self.cali_path+"init_record.txt", "a+")
        
        print (datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        print ("Performing {} calibration for {}".format(source, note))
        print ("Test will assume the temperature is {}".format(settings.temp))
#        self.sbnd.femb_config.initBoard()
#        self.sbnd.femb_config.syncADC([0,1,2,3])
        
        frame_size = 0x07e0
        if (frame_size%13 != 0):
            frame_size = 13 * (frame_size//13)
            
        #Automatic streaming, no trigger
        Trigger_mode = 0
        
        reg31_value = ((Trigger_mode<<31) & 0x80000000) + frame_size
        
        self.sbnd.femb_config.femb.write_reg(31, reg31_value, "wib")
        
        sys.stdout.close()
        sys.stdout = screendisplay
        
    def save_config(self):
        
        with open(self.cali_path + 'configuration.cfg', 'rb') as f:
                data = pickle.load(f)
            
        data['noise'] = self.noise
        data['fpga'] = self.fpga
        data['int'] = self.int
        data['PULSE'] = False
        data['GAIN'] = False
        data['RMS'] = False
        data['ENC'] = False
            
        print ("Save_Config altered {}".format(data))
            
        with open(self.cali_path + 'configuration.cfg', 'wb') as f:
            pickle.dump(data, f, pickle.HIGHEST_PROTOCOL)
            
    def listen_for_trigger_data(self, num_of_packets,
                            num_of_chips,
                            packets_per_file,
                            buffer_size,
                            udp_port,
                            PC_IP1,
                            PC_IP2,
                            PC_IP3,
                            PC_IP4,
                            directory,
                            queue):
    
        self.sbnd.femb_config.femb.get_packets_from_c (num_of_packets,
                                num_of_chips,
                                packets_per_file,
                                buffer_size,
                                udp_port,
                                PC_IP1,
                                PC_IP2,
                                PC_IP3,
                                PC_IP4,
                                directory,
                                queue)
        logging.debug("Done")
        return
    
    def quick_update(self):
        self.sbnd.analyze.UnpackData(path = "default", 
                                 data = "D:\\nEXO\\2018_03_07\\Triggered_200hz\\Separated_Packets\\Chip0_Packet50.dat" , 
                                                       return_data=False)
#        self.sbnd.femb_config.fe_reg.set_fe_board(sts=1,sg=3, sbf = 0, st=2, snc=1,
#                                                          sdacsw1=0, sdacsw2=1, sdac=5)
#        
#        self.sbnd.femb_config.adc_reg.set_adc_board(d=0, pcsr=1, pdsr=1, slp=0, tstin=0,
#                                                    clk = 0, frqc = 0, en_gr = 1, f0 = 0, f1 = 0, 
#                                                    f2 = 0, f3 = 0, f4 = 1, f5 = 0, slsb = 0, show="FALSE")     
#        
#        self.sbnd.femb_config.configAdcAsic(board = "wib")
#        self.sbnd.femb_config.configFeAsic(board = "wib")
        
    def __init__(self):
            self.sbnd = FEMB_DAQ()
            self.cali_path = ""
            self.triggered_path = ""
            
            self.sts_tuple = ((0,"CapOff"), (1,"CapOn"))
            self.snc_tuple = ((0,"900mV"), (1,"200mV"))
            self.sg_tuple  = [(3,"25.0mV")]
            self.st_tuple  = ((2,"0.5us"), (0,"1.0us"), (3,"2.0us"), (1,"3.0us"))
            self.sdc_tuple = ((0, "DC"), (1, "AC"))
            self.sdf_tuple = ((0, "BufOff" ), (1, "BufOnn" )) 
            self.slk0_tuple = ((0,"500pA"), (1, "100pA"))
            self.slk1_tuple = ((0,"pAx10Dis"), (1, "pAx10Enn"))
            
            self.noise      = False
            self.fpga       = False
            self.int        = False
            
if __name__ == "__main__":
#    main().quick_update()
    main().loop()
    
    