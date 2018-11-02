import os
from scripts.raw_process2 import rms_process,ENC_plots,gain_process, baseline_process
from scripts.gain_to_excels import get_gain_results
from scripts.Energy_Analysis import Energy

import openpyxl as px
from openpyxl import Workbook
import glob
import pickle
import gc
import matplotlib as plt

class calibration_analysis:
    def loop(self):
        self.cal_dir_name = (glob.glob(self.cur_path + self.calibration_folder)[0]) + "\\"
            
        print ("\nPath is " + self.cal_dir_name + "\\")

        for self.root, self.dirs, self.files in os.walk(self.cal_dir_name):
            break
        
        #chip_num = int(raw_input( "Enter number of chips analyzed in file\n"))    
        
        print ("Data will be analyzed for {} ASIC chips\n".format(self.chip_num))
        print ("WARNING: Do not run this script in an interactive Python shell like iPython!\n")
        print ("There's a bug in MatPlotLib that will cause a memory leak!\n")
        
        with open(self.cal_dir_name + '\\configuration.cfg', 'rb') as f:
            self.config = pickle.load(f)
#            self.config = 0
#        temp_code = (self.config["temp"])
        
#        if (temp_code == "RT"):
#            print ("Test was done at Room Temperature\n")
#        elif (temp_code == "LN"):
#            print ("Test was done at Liquid Nitrogen Temperature\n")
#        elif (temp_code == "LXe"):
#            print ("Test was done at Liquid Xenon Temperature\n")
#        else:
#            print ("No calibration data found.  Check or run calibration\n")
        
        #Find out what analyses have already been done to suggest the next.
        
        try:
            self.PULSE = self.config["PULSE"]
            self.GAIN = self.config["GAIN"]
            self.RMS = self.config["RMS"]
            self.ENC = self.config["ENC"]
        except:
            pass
#        print (self.config)
        
        self.help_info()
        
        while(1):
        
            if (self.PULSE == 1):
                if (self.GAIN == 1):
                    if (self.RMS == 1):
                        if (self.ENC == 1):
                            print ("It looks like all available analysis was done for this test run\n")
                        else:
                            print ("It looks like the PULSE, GAIN, and RMS analyses have been run.  Run the ENC analysis\n")
                    else:
                        print ("It looks like the PULSE and GAIN analyses have been run.  Run the RMS analysis\n")
                else:
                    print ("It looks like the PULSE analysis has been run.  Run the GAIN analysis\n")
            else:
                print ("It looks like no analyses have been run.  Run the PULSE analysis\n")

            function = input( "Enter 'end, 'help', 'BASE', 'PULSE', 'GAIN', 'RMS', 'ENC', or 'ALL'\n")
            if (function == "end"):
                break
                
            elif (function == "help"):
                self.help_info()
                
            elif (function == "BASE"):
                self.Baseline_analysis()
                
            elif (function == "PULSE"):
                self.PULSE_analysis()
                
            elif (function == "GAIN"):
                self.GAIN_analysis()
                
            elif (function == "RMS"):
                self.RMS_analysis()
                
            elif (function == "ENC"):
                self.ENC_analysis()
				
            elif (function == "Energy"):
                self.Energy_analysis()
                
            elif (function == "ALL"):
                self.Baseline_analysis()
                self.PULSE_analysis()
                self.GAIN_analysis()
                self.RMS_analysis()
#                self.ENC_analysis()
                
            else:
                print ("That's not a function")
                
    def Baseline_analysis(self):
        print ("Beginning Baseline Analysis")
        RMS_wb = Workbook()
        for onedir in self.dirs:
            search_path = self.root + "\\" + onedir + "\\"
            for root_pulse, dirs_pulse, files_pulse in os.walk(search_path):
                break
            
            if (('pedestal.dat' in files_pulse) == True):
                gc.collect()

                RMS_filename = self.root + self.INT_RMS_excel_name
                baseline_process(search_path, RMS_wb)
                RMS_wb.save(filename = RMS_filename)
                
        if (len(RMS_wb.sheetnames) > 1):
            RMS_wb.remove(RMS_wb["Sheet"])
            RMS_wb.save(filename = RMS_filename)

        self.config['BASE'] = 1
        self.BASE = 1
            
        with open(self.cal_dir_name + 'configuration.cfg', 'wb') as f:
                pickle.dump(self.config, f, pickle.HIGHEST_PROTOCOL)
                
        print ("Completed Baseline Analysis")
    
    def PULSE_analysis(self):
        print ("Beginning Pulse Analysis")
        wb_int = Workbook()
#        RMS_filename = self.root + self.INT_RMS_excel_name
#        wb_base = px.load_workbook(RMS_filename)
        
        for onedir in self.dirs:
            search_path = self.root + "\\" + onedir + "\\"
            for root_pulse, dirs_pulse, files_pulse in os.walk(search_path):
                break
            
            if (('cali_intdac' in dirs_pulse) == True):
                gain_process(search_path, wb_int, wb_base=None) 
                int_filename = self.root + self.INT_pulse_excel_name
                wb_int.save(filename = int_filename)

        if (len(wb_int.sheetnames) > 1):
            wb_int.remove(wb_int["Sheet"])
            wb_int.save(filename = int_filename)
            
        self.config['PULSE'] = 1
        self.PULSE = 1
            
        with open(self.cal_dir_name + 'configuration.cfg', 'wb') as f:
                pickle.dump(self.config, f, pickle.HIGHEST_PROTOCOL)
                
        print ("Completed Pulse Analysis")
    
    def GAIN_analysis(self):
        for onedir in self.dirs:
            search_path = self.root + "\\" + onedir + "\\"
            for root_pulse, dirs_pulse, files_pulse in os.walk(search_path):
                break
            
            if (('cali_intdac' in dirs_pulse) == True):
                pulse_filename = self.root + self.INT_pulse_excel_name
                RMS_filename = self.root + self.INT_RMS_excel_name
                get_gain_results(search_path, pulse_filename, RMS_filename,
                                 self.config["temp"])

        self.config['GAIN'] = 1
        self.GAIN = 1
            
        with open(self.cal_dir_name + 'configuration.cfg', 'wb') as f:
                pickle.dump(self.config, f, pickle.HIGHEST_PROTOCOL)
        
    def RMS_analysis(self):
        for onedir in self.dirs:
            search_path = self.root + "\\" + onedir + "\\"
            for root_pulse, dirs_pulse, files_pulse in os.walk(search_path):
                break
            
            if (('pedestal.dat' in files_pulse) == True):
                gc.collect()
                    
                RMS_filename = self.root + self.INT_RMS_excel_name
                
                if (os.path.isfile(RMS_filename)):
                    print ("Internal DAC Pulse")
                    wb = px.load_workbook(RMS_filename)  
                    rms_process(search_path, wb)
                    wb.save(filename = RMS_filename)

        self.config['RMS'] = 1
        self.RMS = 1
            
        with open(self.cal_dir_name + 'configuration.cfg', 'wb') as f:
                pickle.dump(self.config, f, pickle.HIGHEST_PROTOCOL)
    
    def ENC_analysis(self):
        RMS_filename = self.root + self.FPGA_RMS_excel_name
        
        if (os.path.isfile(RMS_filename)):
            print ("External FPGA DAC Pulse")
            wb = px.load_workbook(RMS_filename)  
            plot_dir = self.root +"\\ENC_Plots(FPGA DAC)\\"
            ENC_plots(wb,self.chip_num,plot_dir)
            
        RMS_filename = self.root + self.INT_RMS_excel_name
        
        if (os.path.isfile(RMS_filename)):
            print ("Internal DAC Pulse")
            wb = px.load_workbook(RMS_filename)  
            plot_dir = self.root +"\\ENC_Plots(INT DAC)\\"
            ENC_plots(wb,self.chip_num,plot_dir)
            
        self.config['ENC'] = 1
        self.ENC = 1
            
        with open(self.cal_dir_name + 'configuration.cfg', 'wb') as f:
                pickle.dump(self.config, f, pickle.HIGHEST_PROTOCOL)
				
    def Energy_analysis(self):
        calibration_directory = "D:\\nEXO\\2018_04_18\\Calibration"
        data_directory = "D:\\nEXO\\2018_04_04\\Triggered\\Seperated_Packets\\"
        output_directory = "D:\\nEXO\\2018_04_18"
        samples = 1000
        Energy_data = Energy(cal_directory = calibration_directory, data_directory = data_directory, samples = samples)
		
        for bins in [100,500,1000]:
            fig = plt.figure(figsize=(12,8))
            ax = fig.add_subplot(111)  
            ax.hist(Energy_data,bins=bins)
            ax.set_xlabel("Energy (eV)")
            ax.set_ylabel("Counts")
            ax.set_title("Energy Spectrum")
            fig.savefig ("{}_{}.jpg".format(output_directory, bins))
            fig.clf()
            plt.close()
        
    def help_info(self):
        print ("The analyses in this file are meant to be run after data has been collected from the BNL ASICs.  "
               "The directory above should be in the correct structure.\n")
        
        print ("The PULSE analysis organizes the data taken for each channel/configuration at each test pulse value "
               "into a spreadsheet in the root directory that ends with 'Pulse_Data.xlsx'.\n")
        
        print ("The GAIN analysis will go through the above spreadsheet and use the peak values from the test pulses "
               "to find the gain in ADC counts/injected charge and estimated ADC counts/voltage input for the ADC for "
               "each channel/configuration.  The fitted plots will be in each test directory for the 200 mV baseline data "
               "in the 'test_pulse_fits' folder.  The gain data will be placed in a spreadsheet in the root folder that "
               "ends with 'RMS_Data.xlsx'.\n")
        
        print ("The RMS analysis will import the baseline data taken with no charge input.  It will make a histogram "
               "for each channel/configuration available in the test directory in the 'Histograms' folder.  It will use "
               "the average and RMS baseline data to continue filling in the 'RMS_Data.xlsx' spreadsheet and use the gain "
               "data from the GAIN analysis to find the baseline in mV and the ENC in electrons.  It will do a check for "
               "stuck bits, and mark cells red in the spreadsheet where a stuck bit is detected.\n")
        
        print ("The ENC analysis will use the complete 'RMS_Data.xlsx' spreadsheet to plot the ENC vs. Peaking Time "
               "for each channel at each gain and baseline.  The results of these plots will be in the 'ENC_Plots' folder "
               "visible in the root directory.\n")
        
        print ("Type 'ALL' to do them all sequentially")
        
        print ("Type 'end' to exit.")
        
    def __init__(self):
#        self.cur_path = "Z:\\nEXO - Charge Readout\\Stanford Setup\\Stanford Trip April 2018\\"
        self.cur_path = "D:\\nEXO\\2018_10_19\\"
        self.calibration_folder = "*Calibration_test*"
        self.chip_num = 4
        self.PULSE = None
        self.GAIN = None
        self.RMS = None
        self.ENC = None
        self.config = None
        self.root = None
        self.dirs = None
        self.files = None
        self.cal_dir_name = None
        
        self.FPGA_pulse_excel_name = "\\FPGA_Pulse_Data.xlsx"
        self.INT_pulse_excel_name = "\\Internal_Pulse_Data.xlsx"
        
        self.FPGA_RMS_excel_name = "\\FPGA_RMS_Data.xlsx"
        self.INT_RMS_excel_name = "\\Internal_RMS_Data.xlsx"
            
        
if __name__ == "__main__":
    print ("Start")
    calibration_analysis().loop()

