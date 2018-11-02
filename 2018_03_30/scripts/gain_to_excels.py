# -*- coding: utf-8 -*-
"""
File Name: read_mean.py
Author: GSS
Mail: gao.hillhill@gmail.com
Description: 
Created Time: 3/9/2016 7:12:33 PM
Last modified: 10/13/2016 10:09:29 AM
"""

#defaut setting for scientific caculation
#import numpy
#import scipy
#from numpy import *
#import numpy as np
#import scipy as sp
#import pylab as pl

import openpyxl as px
from openpyxl.styles import Border, Alignment, Font, Side, PatternFill
from openpyxl import Workbook
import numpy as np
import os
from scripts.int_dac_fit import int_dac_fit
from scripts.linear_fit_m import linear_fit
import matplotlib.pyplot as plt
from scripts.raw_process2 import make_title
import sys
import glob
import pickle
import warnings
import copy
from user_settings import user_editable_settings
settings = user_editable_settings()


####################################################################################################################################################
#Loads the existing pulse data spreadsheet and organizes the data for a single sheet into a 3 dimensional matrix
#Returns a 3D matrix 16 channels wide, however many chips long, and the third dimension is the pulse height for a given dac value
def read_gain(filepath,sheetname,dac_steps):

    W = px.load_workbook(filepath)
    p = W[sheetname]
    
    ws_integral = W[sheetname + "_Area"]
    ws_ratio = W[sheetname + "_Ratio"]
    
    dacmean=[]
    integral = []
    ratios = []
    chip_num = settings.chip_num
    for dac_step in range(dac_steps):
        for chip_id in range(chip_num):
            for chn in range(16):
                v1 = p['{}{}'.format(chr(ord('A') + chn + 1), chip_id+4+(dac_step*(chip_num+3)))].value
                dacmean.append(v1)

                v2 = ws_integral['{}{}'.format(chr(ord('A') + chn + 1), chip_id+4+(dac_step*(chip_num+3)))].value
                integral.append(v2)

                v3 = ws_ratio['{}{}'.format(chr(ord('A') + chn + 1), chip_id+4+(dac_step*(chip_num+3)))].value
                ratios.append(v3)
                
    dacmean = np.resize(dacmean, [dac_steps, chip_num,16] )
    integral = np.resize(integral, [dac_steps, chip_num,16] )
    ratios = np.resize(ratios, [dac_steps, chip_num,16] )
    return dacmean,integral,ratios

####################################################################################################################################################
#This function takes an existing plot and some data and does analysis of interest and annotates the existing plot with the results
#It needs the plot, the stats (the list to analyze), a multiplier (to turn the electrons unit into "millions of electrons" for example),
#the subtitle desired, chip number and the written out units to report the slopes in
def plot_stats(plt,stats,chip_id,unit):

#Where on the plot to put these annotations
    average_text = (.65,.25)
    maximum_text = (.65,.20)
    minimum_text = (.65,.15)
    std_text = (.65,.10)
    gain_text = (.65,.05)

#Since the list of slopes can contain more than one chip's worth, this isolates the 16 slopes we care about for this chip plot and makes
#it an array for easier analysis below.  It also multiplies the numbers right away so they're in the format we want
    slope_stat = np.array(stats[16*(chip_id)+0:16*(chip_id)+16])



    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=RuntimeWarning)
        average = str(round(np.nanmean(slope_stat),5))
    maximum = str(round(max(slope_stat),5))
    minimum = str(round(min(slope_stat),5))
    std_dev = str(round(np.std(slope_stat),5))

    plt.annotate('Average gain is '+average, xy=average_text,  xycoords='axes fraction',
    xytext=average_text, textcoords='axes fraction',
    horizontalalignment='left', verticalalignment='center',
    )

    plt.annotate("Maximum gain is "+maximum, xy=maximum_text,  xycoords='axes fraction',
    xytext=maximum_text, textcoords='axes fraction',
    horizontalalignment='left', verticalalignment='center',
    )

    plt.annotate("Minimum gain is "+minimum, xy=minimum_text,  xycoords='axes fraction',
    xytext=minimum_text, textcoords='axes fraction',
    horizontalalignment='left', verticalalignment='center',
    )

    plt.annotate("Standard Deviation of gains is "+std_dev, xy=std_text,  xycoords='axes fraction',
    xytext=std_text, textcoords='axes fraction',
    horizontalalignment='left', verticalalignment='center',
    )

    plt.annotate("Gain units in ADC counts per "+unit, xy=gain_text,  xycoords='axes fraction',
    xytext=gain_text, textcoords='axes fraction',
    horizontalalignment='left', verticalalignment='center',
    )

####################################################################################################################################################
#Opens an existing pulse data spreadsheet, and for every chip/channel/configuration, analyzes the increase in pulse height
#for increasing DAC values.  It fits this line and plots all the data in the given directories.  It also creates a spreadsheet
#for later RMS/ENC analysis, and populates it with the gain data, so the RMS and mean in ADC counts can be converted to electrons or mV
def get_gain_results(test_dir, pulse_filepath, rms_filepath, temp):


#Utilize existing knowledge on what each DAC step corresponds to in mV.  The below file must be in the same directory as the .py file
    ln2_int_vlt_slope = int_dac_fit(1)
    rt_int_vlt_slope = int_dac_fit(0)

#By looking at the name of the spreadsheet, determine if this test was done by the FPGA or internal DAC.  These distinctions are important
#because 'one DAC step' means something different for the FPGA vs. internal DAC, which is different at Liquid Nitrogen and Room Temperature
    if (temp == "RT"):
        vlt_slope = rt_int_vlt_slope
    elif (temp == "LN" or temp == "LXe"):
        vlt_slope = ln2_int_vlt_slope
    else:
        sys.exit("Wrong file name")
        
    print ("Internal DAC Pulse")
    dac_steps = 64
    plot_folder_name = "\\int_pulse_fits\\" 
#Get the basics about the test directory we're in

    config_file = (glob.glob(test_dir + "*chip_settings*")[0])
        
    with open(config_file, 'rb') as f:
        data = pickle.load(f)
        gain = (data["gain"])
        peak = (data["peak"])
        
    sheet_title = gain[0:4] + "," + peak[0:3]
    print (sheet_title)

    
#Create a new workbook for RMS/ENC data
    if (os.path.isfile(rms_filepath)):
        RMS_wb = px.load_workbook(rms_filepath)
#        test_titles = [sheet_title+",200", sheet_title+",900"]
#        test_titles = [sheet_title+",200"]
#        for title in test_titles:
#            if (title in RMS_wb.sheetnames):
#                RMS_wb.remove_sheet(RMS_wb[title])   
        RMS_ws1 = RMS_wb.active
    else:
        RMS_wb = Workbook()
        RMS_ws1 = RMS_wb.create_sheet()
#       RMS_ws2 = RMS_wb.create_sheet()
 
#Now that we know which directory, we can create the folder there for plots
    result_path= test_dir + plot_folder_name

    try: 
        os.makedirs(result_path)
    except OSError:
        if os.path.exists(result_path):
            pass


#This function takes in the filepath to the pulse sheet, the sheetname you want analyzed, total number of chips and DAC steps
#(which changes if using FPGA or Interna DAC) and returns the entire dataset for the sheet, organized in a 3D matrix
    chip_num = settings.chip_num
    dacmean, integral, ratios = read_gain(pulse_filepath,sheet_title,dac_steps)

    peak_slope = []
    integral_slope = []
    ratio_slope = []

    for chip_id in range(chip_num):

#Each chip gets its own plot.  Because you're switching back and forth between two plots, I found this to be the easiest way
        plt.figure(0, figsize=(12,8))
        electron_plot = plt.subplot(1,1,1)
        electron_plot.figsize=(9,6)
        electron_plot.set_xlabel("Test Charge Injection (fC)")
#So there's not so many zeroes on the electron plt
#        electron_plot.ticklabel_format(axis='x', style='sci', scilimits=(-2,2))
        electron_plot.set_ylabel("ADC counts")
        electron_plot.set_title("Chip {} ADC Output for various injected charges".format(chip_id))
        
        plt.figure(1, figsize=(12,8))
        integral_plot = plt.subplot(1,1,1)
        integral_plot.figsize=(9,6)
        integral_plot.set_xlabel("Test Charge Injection (fC)")
#So there's not so many zeroes on the electron plt
#        integral_plot.ticklabel_format(axis='x', style='sci', scilimits=(-2,2))
        integral_plot.set_ylabel("ADC counts * Time")
        integral_plot.set_title("Chip {} ADC Pulse Integral for various injected charges".format(chip_id))

        plt.figure(2, figsize=(12,8))
        ratio_plot = plt.subplot(1,1,1)
        ratio_plot.figsize=(9,6)
        ratio_plot.set_xlabel("Test Charge Injection (fC)")
#So there's not so many zeroes on the electron plt
#        ratio_plot.ticklabel_format(axis='x', style='sci', scilimits=(-2,2))
        ratio_plot.set_ylabel("mV / (ADC counts * Time)")
        ratio_plot.set_title("Chip {} Peak/Integral Ratio for various injected charges".format(chip_id))
#Name of the file to save it as
        electron_path = result_path + "electrons_chip{}".format(chip_id)
        integral_path = result_path + "integral_chip{}".format(chip_id)
        ratio_path = result_path + "ratio_chip{}".format(chip_id)
        
        for chn in range(16):

#This returns the progressively increasing pulse height for a given channel in ADC counts
#It's a one dimensional list with the same alements as there were DAC steps, starts at DAC step 1
#Because DAC step 0 means no pulse.
            adc_np = dacmean[1:,chip_id,chn]
            int_np = integral[1:,chip_id,chn]
            rat_np = ratios[1:,chip_id,chn]

#This function does all the linear analysis and will return the slope, y-intercept, x-values, and ADC counts (unchanged from when
#passed to the function).  It takes in the chip and channel, which come from the for loop, the adc count which acts as the y-values
#as obtained in the previous line, the slope that turns "DAC number" into the desired x-value.  The 1 means you get back the injected
#charge, in which case the "gain" parameter is meaningless.
            a,econstant, adc_counts, fC1 = linear_fit(chip_id, chn, adc_np,vlt_slope)
            peak_slope.append(a)
            electron_plot.scatter(fC1, adc_counts, marker='.')
            electron_plot.plot(fC1, adc_counts)
            
            b,econstant, int_counts, fC2 = linear_fit(chip_id, chn, int_np,vlt_slope)
            integral_slope.append(b)
            integral_plot.scatter(fC2, int_counts, marker='.')
            integral_plot.plot(fC2, int_counts)
            
            units = []
            for i in (range(len(rat_np))):
                units.append(vlt_slope * i * 185)
            valid_min = 3
            valid_max  = 11
            y = copy.deepcopy(rat_np[valid_min:valid_max])
            x = copy.deepcopy(units[valid_min:valid_max])
            av = np.nanmean(y, dtype=float)
            
            ratio_slope.append(av)
            ratio_plot.scatter(x, y, marker='.')
            ratio_plot.plot(x, y)
    
        plot_stats(electron_plot,peak_slope,chip_id,"fC")
        plot_stats(integral_plot,integral_slope,chip_id,"fC")
        plot_stats(ratio_plot,ratio_slope,chip_id,"fC")
        
#This is just for the electron plot, it looks weird with a tick that says negative electrons
        xticks = electron_plot.xaxis.get_major_ticks()
        xticks[0].label1.set_visible(False)
        xticks = integral_plot.xaxis.get_major_ticks()
        xticks[0].label1.set_visible(False)
        xticks = ratio_plot.xaxis.get_major_ticks()
        xticks[0].label1.set_visible(False)

#Now that the plot is done, it's selected, saved and cleared, because with this figure(#) method, the next iteration wont create a new plot
        plt.figure(0)
        plt.savefig (electron_path+".jpg")
        plt.clf()
        
        plt.figure(1)
        plt.savefig (integral_path+".jpg")
        plt.clf()
        
        plt.figure(2)
        plt.savefig (ratio_path+".jpg")
        plt.clf()
        
        plt.close(0)
        plt.close(1)
        plt.close(2)

#After all the plots are done, it's time to create the new spreadsheet for ENC analysis  First, the collection of each channel's gain is organized
    peak_slope = np.resize(peak_slope, [chip_num,16])
    integral_slope = np.resize(integral_slope, [chip_num,16])
    ratio_slope = np.resize(ratio_slope, [chip_num,16])
    RMS_ws1.title = sheet_title+",200"
#    RMS_ws2.title = sheet_title+",900"

    make_title(3,chip_num,"Gain in ADC counts/fC",RMS_ws1)
    make_title(4,chip_num,"Gain in integral/fC",RMS_ws1)
    make_title(5,chip_num,"Average in ratio/fC",RMS_ws1)
#    make_title(4,chip_num,"Gain in ADC counts/electron",RMS_ws2)
   
#For each chip/channel on both 200 and 900 sheets, it just places the gain data.  Format for that sheet is always 
#chip_id+4+(chunk*(chip_num+3)) where the chunk should match up with the ones titled above

    for chip_id in range(chip_num):
        for chn in range(16):

                RMS_ws1['{}{}'.format(chr(ord('A') + chn + 1), 
                        chip_id+4+(3*(chip_num+3)))].value = peak_slope[chip_id][chn]
#                RMS_ws2['{}{}'.format(chr(ord('A') + chn + 1), 
#                        chip_id+4+(4*(chip_num+3)))].value = eslope[chip_id][chn]
    
                RMS_ws1['{}{}'.format(chr(ord('A') + chn + 1), 
                        chip_id+4+(4*(chip_num+3)))].value = integral_slope[chip_id][chn]
                        
                RMS_ws1['{}{}'.format(chr(ord('A') + chn + 1), 
                        chip_id+4+(5*(chip_num+3)))].value = ratio_slope[chip_id][chn]

#Save after every sheet, which is useful when debugging
        RMS_wb.save(filename = rms_filepath)

#After every sheet is done, delete that first empty sheet and save
    if ("Sheet" in RMS_wb.sheetnames):
        RMS_wb.remove_sheet(RMS_wb["Sheet"])   
    RMS_wb.save(filename = rms_filepath)


#Part of the program that runs, change filenames and directories and chip numbers as needed
#root_path = "D:\\femb_3\\2016_10_07"
#pulse_filename = "RF_Pulse_Data.xlsx"
#rms_filename = "R_RMS_Data.xlsx"
#chip_num = 4
#get_gain_results(root_path,pulse_filename,rms_filename,chip_num)


