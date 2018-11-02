# -*- coding: utf-8 -*-
"""
File Name: init_femb.py
Author: GSS
Mail: gao.hillhill@gmail.com
Description: 
Created Time: 7/15/2016 11:47:39 AM
Last modified: 10/18/2016 4:37:37 PM
"""

#defaut setting for scientific caculation
#import numpy
#import scipy
#from numpy import *
#import numpy as np
#import scipy as sp
#import pylab as pl
from scipy.integrate import simps
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Font, Side, PatternFill
import numpy as np
import struct
import os
from scripts.detect_peaks import detect_peaks
import matplotlib
#matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy import stats
import sys
import glob
import pickle
import re
from scripts.Data_Analysis import Data_Analysis
import glob
import pickle
import warnings
import gc
from user_settings import user_editable_settings
settings = user_editable_settings()
analyze = Data_Analysis()

####################################################################################################################################################


def baseline_process(test_dir, wb):
    
    rms_data_file = test_dir + "\\pedestal.dat"

#Name the sheet based on the directory you're looking into for the above file
    config_file = (glob.glob(test_dir + "*chip_settings*")[0])
    with open(config_file, 'rb') as f:
        data = pickle.load(f)
        baseline = (data["base"])
        gain = (data["gain"])
        peak = (data["peak"])
        
    sheet_title = gain[0:4] + "," + peak[0:3] + "," + baseline[0:3]
    print (sheet_title)

    with open(rms_data_file, 'rb') as f:
        raw_data = f.read()
        
    start_of_packet = (b"\xde\xad\xbe\xef")
    start_of_chip = []
    for m in re.finditer(start_of_packet, raw_data):
        start_of_chip.append(m.start())
        
    chip_num = settings.chip_num
    if (len(start_of_chip) != chip_num):
        print ("RMS Analysis--> {} doesn't have {} chips in the file!".format(rms_data_file, chip_num))
        
    separated_by_chip = [(raw_data[start_of_chip[0] : start_of_chip[1]]),
                         (raw_data[start_of_chip[1] : start_of_chip[2]]),
                         (raw_data[start_of_chip[2] : start_of_chip[3]]),
                         (raw_data[start_of_chip[3] : ])]
        
    
    chip_data = [[],[],[],[]]
    for i in range(chip_num):
        chip_data[i] = analyze.UnpackData(path = "bytes", data = separated_by_chip[i], return_data = True)


#Histograms are plotted in the folder for each configuration
    plot_dir = test_dir + "\\Histograms\\"

    try: 
        os.makedirs(plot_dir)
    except OSError:
        if os.path.exists(plot_dir):
            pass
#Mean, Standard Deviation and stuck bit arrays that get passed to the final Excel spreadsheet
    std_np = []
    mean_np = []
    stuck_np = []

#Positioning for histogram text
    av_text = (.8,-.14)
    mode_text = (.8,-.19)
    std_text = (-.10,-.24)
    samples_text = (-.10,-.14)
    samples_within_text = (-.10,-.19)
    st_text = (.8,-.24)
    stuck_text = (.4,-.24)

    for chip in range(chip_num):
        for chn in range(16):
#For a given chip and channel's data, read and find the following parameters
            print ("Chip {}, chn {}".format(chip, chn))
            np_data_all = np.array(chip_data[chip][chn])
            np_data = []
            for i in (np_data_all):
                mod = i % 64
                if (mod != 0 and mod!=1 and mod!=63):
                    np_data.append(i)
            if (len(np_data) <2):
                np_data = [0,0]
            datamean = np.mean(np_data)
            mean_np.append (datamean)
            std = np.std(np_data)
            std_np.append (std)
            mode = stats.mstats.mode(np_data)
            total_num = len(np_data)
            
            low_range = datamean - (3*std)
            high_range = datamean + (3*std)

            num_within = 0

#As part of stuck bit check, finds the amount of samples within 3 sigma

            for i in range(total_num):
                if (np_data[i] > low_range) and (np_data[i] < high_range):
                    num_within += 1
             
            maximum=max(np_data)
            minimum=min(np_data)

            plot_path = plot_dir + "Chip" + str(chip+1) + "_Ch" + str(chn)

            bins = maximum - minimum
            if (bins < 1):
                bins = 1
            if (bins >100):
                bins = 100

            fig = plt.figure(figsize=(12,8))
            ax = fig.add_subplot(111)  
            ax.hist(np_data,bins=bins)
            ax.set_xlabel("ADC Counts")
            ax.set_ylabel("Occurences")
            ax.set_title("ADC Count Distribution for Chip "+str(chip+1)+", Channel "+str(chn))
# Shrink current axis's height by 10% on the bottom
            box = ax.get_position()
            ax.set_position([box.x0, box.y0 + box.height * 0.1,
            box.width, box.height * 0.9])
#Self explanatory, it takes the statistics found above and plots them in the designated place.  'axes fraction' lets you give the coordinates
#in terms of the axes, where 0 is one extreme and 1 is the other.  Without that, the coordinates would be in data format, which is harder to use

            ax.annotate("Average = "+str(round(datamean,2)), xy=av_text,  xycoords='axes fraction',
            xytext=av_text, textcoords='axes fraction',
            horizontalalignment='left', verticalalignment='left',
            )

            ax.annotate("Mode = "+str(int(mode[0][0])), xy=mode_text,  xycoords='axes fraction',
            xytext=mode_text, textcoords='axes fraction',
            horizontalalignment='left', verticalalignment='left',
            )

            ax.annotate("Standard deviation = "+str(round(std,2)), xy=std_text,  xycoords='axes fraction',
            xytext=std_text, textcoords='axes fraction',
            horizontalalignment='left', verticalalignment='left',
            )

            ax.annotate("Total samples = "+str(total_num), xy=samples_text,  xycoords='axes fraction',
            xytext=samples_text, textcoords='axes fraction',
            horizontalalignment='left', verticalalignment='left',
            )

            ax.annotate("Total samples in 3 sigma = "+str(num_within), xy=samples_within_text,  xycoords='axes fraction',
            xytext=samples_within_text, textcoords='axes fraction',
            horizontalalignment='left', verticalalignment='left',
            )

            ax.annotate("S/T Ratio = "+str(round(float(num_within)/float(total_num),6)), xy=st_text,  xycoords='axes fraction',
            xytext=st_text, textcoords='axes fraction',
            horizontalalignment='left', verticalalignment='left',
            )

            r64 =  int(mode[0][0])%64
            if (r64>3) and (r64<62) and ((float(num_within)/float(total_num))>0.995):
                stuck_np.append(1)
            else: 
                stuck_np.append(0)
                plt.annotate("Flagged as Stuck", xy=stuck_text,  xycoords='axes fraction',
                xytext=stuck_text, textcoords='axes fraction',
                horizontalalignment='left', verticalalignment='left', color='red',
                )
            fig.savefig (plot_path+".jpg")
            fig.clf()
            plt.close()
            gc.collect()

#Does the stuck bit analysis, finding out whether the mean/mode is close to a multiple of 64, which is indicative of a stuck bit
#A 1 means it looks good, a 0 means it could be a stuck bit

    std_np = np.resize(std_np,[chip_num,16])
    mean_np = np.resize(mean_np,[chip_num,16]) 
    stuck_np =  np.resize(stuck_np,[chip_num,16])
    
    ws = wb.create_sheet(0)
    ws.title = "{},{},{}".format(baseline,gain,peak)

    center = Alignment(horizontal='center')

    title = ("Baseline = {}, Gain = {}, Peaking Time = {}".format(baseline, gain, peak))
#First two rows are for title blocks
    ws.merge_cells('B1:Q1')
    ws['B1'].value = title
    ws['B1'].alignment = center

#Make the formatting for each new chunk of data (The Chip# and Channel# cells as well as the merged title.  Send the number of the chunk (0, 1, 2),
#the number of chips being analyzed, the title you want, and ws.  These must match up with the actual data bring put in these chunks below
    make_title(0,4,"Mean Value (ADC Counts)",ws)
    make_title(1,4,"RMS Width (ADC Counts)",ws)
    make_title(2,4,"Stuck Bit Matrix - A 0 indicates that a stuck bit was detected for that channel",ws)
    make_title(5,4,"Mean value in mV",ws)
    make_title(6,4,"RMS in electrons",ws)
    
    for chip_id in range(chip_num):
        for chn in range(16):
#            if ((mv_slope[chip_id,chn] != None) and (electron_slope[chip_id,chn])):
            ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4)].value=mean_np[chip_id,chn]
            ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4+(chip_num+3))].value=std_np[chip_id,chn]
            ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4+(2*(chip_num+3)))].value=stuck_np[chip_id,chn]

#Choose which chunks of data you want to mark as stuck or not.  You need to pass the stuck matrix, total chips, which chunk to analyze and ws
    for chunk in range(3):
        mark_stuck_bits(stuck_np,chip_num,chunk,ws)

    for chunk in range(6,7):
        mark_stuck_bits(stuck_np,chip_num,chunk,ws)
        
    raw_data = []
    start_of_chip = []
    separated_by_chip = []
    chip_data = []
    std_np = []
    mean_np = []
    stuck_np = []
    
def rms_process(test_dir, wb):
    ws = wb.active
    #Get pulsed spreadsheet
    
#Grab the existing slope data for each channel as found by the previous script and put it in a matrix to use
    std_slope = []
    electron_slope = []

    for chip_id in range(settings.chip_num):
        for chn in range(16):
            std_slope.append(ws['{}{}'.format(chr(ord('A') + chn + 1), chip_id+4+(settings.chip_num+3))].value)
        
            electron_slope.append(ws['{}{}'.format(chr(ord('A') + chn + 1), chip_id+4+(3*(settings.chip_num+3)))].value)

    std_slope = np.resize(std_slope, [settings.chip_num,16])
    electron_slope = np.resize(electron_slope, [settings.chip_num,16])

#    Couldn't get borders to work, maybe at a later time
#    ws.cell('B2').border = Border(outline=Side(border_style='thick'))
    
    
    #Get RMS Spreadsheet
    rms_values = []
    for chip_id in range(settings.chip_num):
        for chn in range(16):
#                ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4)].value=mean_np[chip_id,chn]
#                ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4+(settings.chip_num+3))].value=std_np[chip_id,chn]
#                ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4+(2*(settings.chip_num+3)))].value=stuck_np[chip_id,chn]
#    #Sixth chunk is Mean value in mV (baseline)
#                ws['{}{}'.format(chr(ord('A') + chn + 1),
#                                 chip_id+4+(5*(chip_num+3)))].value=mean_np[chip_id,chn]/mv_slope[chip_id,chn]
#Seventh chunk is Standard Deviation in electrons (ENC)
            ws['{}{}'.format(chr(ord('A') + chn + 1),
                             chip_id+4+(6*(settings.chip_num+3)))].value=(std_slope[chip_id,chn]/electron_slope[chip_id,chn])/0.0001602176
            rms_values.append((std_slope[chip_id,chn]/electron_slope[chip_id,chn])/0.0001602176)
               #fC per electron
               
    chns = []
    for i in range(len(rms_values)):
        chns.append(i)
    fig = plt.figure(figsize=(16, 12), dpi=80)
    ax = fig.add_subplot(1,1,1)
    ax.scatter(chns, rms_values)
    ax.set_xlabel('Channels', fontsize = 24)
    ax.set_ylabel('Noise (RMS Electrons)', fontsize = 24)
    ax.set_title("RMS Noise for each channel", fontsize=30)
    ax.set_xlim([-1,65])
    ax.tick_params(labelsize=18)
    fig.savefig(test_dir + "Noise.jpg")
    plt.close(fig)
    

####################################################################################################################################################

def gain_process(directory, wb, wb_base):
    
#    ws_base = wb_base.active
#    baselines = [[],[],[],[]]
#    for chip in range(settings.chip_num):
#        for chn in range(16):
#            baselines[chip].append(ws_base.cell(row=4 + chip, column=2 + chn).value)
    
    ws = wb.active
    ws = wb.create_sheet(0)
    ws_area = wb.create_sheet(0)
    ws_ratio = wb.create_sheet(0)

    calidir ="cali_intdac\\"
    init = "intdac_"   
    print ("Internal DAC Pulse")
    print (directory)

    config_file = (glob.glob(directory + "*chip_settings*")[0])
    
    with open(config_file, 'rb') as f:
        data = pickle.load(f)
    baseline = (data["base"])
    gain = (data["gain"])
    peak = (data["peak"])
    print("Peak is {}".format(peak))
    if (peak == "0.5us"):
        calibration_bounds = [26,35]
    elif (peak == "1.0us"):
        calibration_bounds = [26,35]
    elif (peak == "2.0us"):
        calibration_bounds = [24,39]
    elif (peak == "3.0us"):
        calibration_bounds = [20,45]
    else:
        sys.exit("No peak found")

#Build the title of the sheet
    sheet_title = gain[0:4] + "," + peak[0:3]
    print (sheet_title)
    ws.title = sheet_title
    ws_area.title = sheet_title + "_Area"
    ws_ratio.title = sheet_title+ "_Ratio"
    done = False
    
    center = Alignment(horizontal='center')
    for sheets in [ws, ws_area, ws_ratio]:
        sheets.merge_cells('B1:Q1')
        sheets['B1'].alignment = center
    ws['B1'].value = "Average pulse height corrected for baseline (ADC counts)"
    ws_area['B1'].value = "Average area under curve corrected for baseline (ADC counts * time)"
    ws_ratio['B1'].value = "Ratio between pulse height and area"
#    for dacvalue in range(0, 14, 1):
    for dacvalue in range(0, 14, 1):
        print("Collecting data for DAC step {}".format(dacvalue))
        
        plot_directory = directory + "cali_intdac\\Amp{}\\".format(dacvalue)
        try: 
            os.makedirs(plot_directory)
        except OSError:
            if os.path.exists(plot_directory):
                pass
            

#Find the total length of raw data file in bytes and read it into memory
        raw_data_file = directory + calidir + init + "%x"%dacvalue + ".dat"

        with open(raw_data_file, 'rb') as f:
            raw_data = f.read()

        test0 = (b"\xde\xad\xbe\xef\xca\xfe[\x00-\xff][\x00-\xff][\x00-\xff][\x00-\xff][\x00-\xff][\x00-\xff]" +
                b"\x00")
        
        test1 = (b"\xde\xad\xbe\xef\xca\xfe[\x00-\xff][\x00-\xff][\x00-\xff][\x00-\xff][\x00-\xff][\x00-\xff]" +
                b"\x01")
        
        test2 = (b"\xde\xad\xbe\xef\xca\xfe[\x00-\xff][\x00-\xff][\x00-\xff][\x00-\xff][\x00-\xff][\x00-\xff]" +
                b"\x02")
        
        test3 = (b"\xde\xad\xbe\xef\xca\xfe[\x00-\xff][\x00-\xff][\x00-\xff][\x00-\xff][\x00-\xff][\x00-\xff]" +
                b"\x03")
        test = [test0, test1, test2, test3]
        all_packets = [[],[],[],[]]
        for chip in [0,1,2,3]:
            count = 0
            for m in re.finditer(test[chip], raw_data):
    #            print (m)
    #            print (m.start())

    #                print ("Chip {} starts at {}".format(chip, m.start()))
                count = count+1
                all_packets[chip].append(m.start())
    #            sys.exit("first one")

        chip_num = settings.chip_num    

        
        chip_data = [[],[],[],[]]
        peak_data = [[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                     [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                     [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                     [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],]
        integral_data = [[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                         [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                         [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                         [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],]
        for i in range(4):
            for j in range(len(all_packets[i])):
                all_packets[i][j] = int(all_packets[i][j]/2)
        
        print("Looking into each chip")
#        print(all_packets[0])
#        print(all_packets[1])
        for i in [0,1,2,3]:
            chip_data[i] = analyze.UnpackDataPulses(data = raw_data, 
                     all_starts = all_packets[i], return_data = True)
            print("Data Analyzed for chip {}!".format(i))
            fig = analyze.UnpackDataPulses(data = raw_data,
                     all_starts = all_packets[i], return_data = False, plot_length = 1000)
            print("Plot Created for chip {}!".format(i))
            fig.savefig (plot_directory + "Chip{}Amp{}".format(i,dacvalue)+".jpg")
            print("Figure Saved for chip {}!".format(i))
            plt.close(fig)
            print("Figure Closed for chip {}!".format(i))
#        print_pulses(plot_dir, chip_data, value_name)
        for sheets in [ws, ws_area, ws_ratio]:
    
    #Current way of doing the make_title function, but with a title that changes with each DAC step
            row1=2+(dacvalue*(chip_num+3))
            sheets.merge_cells(start_row=row1,start_column=2,end_row=row1,end_column=17)
            sheets['{}{}'.format("A",row1)].value = str(dacvalue)

        for chip in [0,1,2,3]:
            print("Analyzing chip {}".format(chip))
            for chn in range(16):
                print("Analyzing channel {}".format(chn))

#Pull out one channel's array, get 
                np_data = np.array(chip_data[chip][chn])
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore", category=RuntimeWarning)
                    pedmean = np.nanmean(np_data[0:2048])
                maxmean = np.max(np_data[0:2048])
                
#                print("Data for {} {} is {}".format(chip,chn,np_data[0:2048]))

#Find peaks while inputting the "minimum peak height", which must be higher than half the max value found 
#and "minimum peak distance".  Returns the index of the peaks
                peaks_index = detect_peaks(x=np_data, mph=((maxmean + pedmean) / 2), mpd=(500/2)) 
                peaks_value = []
                pulses = 0
                index = 1
                integral = []
                
#                if ((dacvalue < 14) and (chip == 3) and (chn == 10)):
#                    print ("Chip {}, Channel {}".format(chip, chn))
                while(pulses < 10):
                    corrected_data = []
                    if (index > (len(peaks_index) - 1)):
                        break
                    #15 us before and after the peak
                    begin = peaks_index[index] - 30
                    end = peaks_index[index] + 30
                                     
                    #data outside of pulse to get representative baseline
                    begin_data = np_data[begin:begin+15]
                    end_data = np_data[end-15:end]
                    baseline = np.concatenate((begin_data,end_data))
                    outside_baseline = np.nanmean(baseline)
                    
                    for i in range(len(np_data)):
                       corrected_data.append(np_data[i] - outside_baseline)
                               
#                    np_data[i] = np_data[i] - baselines[chip][chn]
                    
                    peak_only = corrected_data[begin:end]
                    
#                        fig, ax = analyze.quickPlot(peak_data)
#                        plt.show()
#                        resp = raw_input("Plot ok? y/n\n")
#                        if (resp != "y"):
#                            index = index + 1
#                            continue

                    integral_window = peak_only[calibration_bounds[0]:calibration_bounds[1]]
#                            area = simps(y=integral_data2, dx = 5E-7)
                    total_sum = 0
                    for point in integral_window:
                        total_sum = total_sum + point
                    integral.append(total_sum)
                    subtitle_text = (.25,.94)
                    fig, ax = analyze.quickPlot(peak_only)
                    
                    for xc in calibration_bounds:
                        plt.axvline(x=xc/2, color='r', linestyle='--')
                    ax.set_title("Chip{}Chn{}Amp{}_{}".format(chip,chn,dacvalue,pulses), fontsize=30)
                    ax.xaxis.label.set_size(30)
                    ax.yaxis.label.set_size(30)
                    for tick in ax.xaxis.get_major_ticks():
                        tick.label.set_fontsize(20)
                    for tick in ax.yaxis.get_major_ticks():
                        tick.label.set_fontsize(20) 
                    ax.annotate("Sum is {}".format(total_sum), xy=subtitle_text,  xycoords='axes fraction',
                        xytext=subtitle_text, textcoords='axes fraction',
                        horizontalalignment='center', verticalalignment='center', fontsize = 20
                        )
                    fig.savefig (plot_directory + "Chip{}Chn{}Amp{}_{}".format(chip,chn,dacvalue,pulses)+".jpg")
                    plt.close(fig)
                    
                    index = index + 1
                    pulses = pulses + 1

                with warnings.catch_warnings():
                    warnings.simplefilter("ignore", category=RuntimeWarning)
                    average_integral = np.mean(np.array(integral))
                    integral_data[chip][chn] = (average_integral)

                for i in peaks_index :
                    peaks_value.append(np_data[i])

                if len(peaks_value) != 0 :
                    print("There were peaks")
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore", category=RuntimeWarning)
                        peaksmean = np.nanmean(peaks_value[0:-1])
                        peak_data[chip][chn] = peaksmean
#                        print ("Peaks_value is {}".format(peaks_value))
#                        print ("Peak data {} {} is {}".format(chip, chn, peaksmean))
                else:
                    print("There were no peaks")
                    peaksmean = pedmean
                    peak_data[chip][chn] = peaksmean
                print("And the average is {:.2f}".format(peak_data[chip][chn]))  
        ft = Font(bold=True)
        print("Writing Excel File")
#        print (peak_data)
#        print (integral_data)
        for chip_id in [0,1,2,3]:
            for chn in range(16):
                if (chip_id == 0):

                    vl = "Channel " + str(chn) 
                    ws['{}{}'.format(chr(ord('A') + chn + 1), 3+(dacvalue*(chip_num+3)))].value = vl
                    ws['{}{}'.format(chr(ord('A') + chn + 1), 3+(dacvalue*(chip_num+3)))].font = ft
                       
                    ws_area['{}{}'.format(chr(ord('A') + chn + 1), 3+(dacvalue*(chip_num+3)))].value = vl
                    ws_area['{}{}'.format(chr(ord('A') + chn + 1), 3+(dacvalue*(chip_num+3)))].font = ft
                            
                    ws_ratio['{}{}'.format(chr(ord('A') + chn + 1), 3+(dacvalue*(chip_num+3)))].value = vl
                    ws_ratio['{}{}'.format(chr(ord('A') + chn + 1), 3+(dacvalue*(chip_num+3)))].font = ft
                             
                             
                             
                column = chr(ord('A') + chn + 1)
                row = chip_id+4+(dacvalue*(chip_num+3))
                ws['{}{}'.format(column, row)].value = peak_data[chip_id][chn]
#                print ("In {}{}, we printed {}".format(column, row, peak_data[chip_id][chn]))
                
                ws_area['{}{}'.format(column, row)].value = integral_data[chip_id][chn]
#                print ("In {}{}, we printed {}".format(column, row, integral_data[chip_id][chn]))
                
                if (integral_data[chip_id][chn] != 0):
                    result = (peak_data[chip_id][chn])/(integral_data[chip_id][chn])
                    ws_ratio['{}{}'.format(chr(ord('A') + chn + 1), chip_id+4+(dacvalue*(chip_num+3)))].value = result
                else:
                    ws_ratio['{}{}'.format(chr(ord('A') + chn + 1), chip_id+4+(dacvalue*(chip_num+3)))].value = 0

            ws['{}{}'.format("A", chip_id+4+(dacvalue*(chip_num+3)))].value = "Chip " + str(chip_id)
            ws['{}{}'.format("A", chip_id+4+(dacvalue*(chip_num+3)))].font = ft
               
            ws_area['{}{}'.format("A", chip_id+4+(dacvalue*(chip_num+3)))].value = "Chip " + str(chip_id)
            ws_area['{}{}'.format("A", chip_id+4+(dacvalue*(chip_num+3)))].font = ft
                    
            ws_ratio['{}{}'.format("A", chip_id+4+(dacvalue*(chip_num+3)))].value = "Chip " + str(chip_id)
            ws_ratio['{}{}'.format("A", chip_id+4+(dacvalue*(chip_num+3)))].font = ft
        print ("DAC Step: %x, Average Amplitude for Chip 0, Channel 0: %d"%(dacvalue, peak_data[0][0]))
        if (done == True):
            return

#Make the formatting for each new chunk of data (The Chip# and Channel# cells as well as the merged title.  Takes in the number of the chunk (0, 1, 2),
#the number of chips being analyzed, the title you want, and ws.
def make_title(num,chip_num,title,ws):
        center = Alignment(horizontal='center')
        ft = Font(bold=True)
        title_row=2+(num*(chip_num+3))
        ws.merge_cells(start_row=title_row,start_column=2,end_row=title_row,end_column=17)
        ws['{}{}'.format("B", title_row)].value = title
        ws['{}{}'.format("B", title_row)].alignment = center

        for chip_id in range(chip_num):
#Adds the "Chip #" Field
            ws['{}{}'.format("A", 4+chip_id+(num*(chip_num+3)))].value = "Chip " + str(chip_id)
            ws['{}{}'.format("A", 4+chip_id+(num*(chip_num+3)))].font = ft

        for chn in range(16):
#Adds the "Channel #" field            
            ws['{}{}'.format(chr(ord('A') + chn + 1),3+(num*(chip_num+3)))].value = "Channel " + str(chn)
            ws['{}{}'.format(chr(ord('A') + chn + 1),3+(num*(chip_num+3)))].font = ft

#Marks each data point as stuck or not.  You need to pass the stuck matrix, total chips, which chunk to analyze and ws
def mark_stuck_bits(stuck_np,chip_num,chunk,ws):

    redFill = PatternFill(start_color='FFC7CE', end_color='CEC7FF', fill_type='solid')
    redFont = Font(color='9C0006')

    for chip_id in range(chip_num):
        for chn in range(16):

#Makes the cells red if they represent a channel with stuck bits
            stuck = stuck_np[chip_id,chn]
            if (stuck == 0):
                ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4+(chunk*(chip_num+3)))].fill=redFill
                ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4+(chunk*(chip_num+3)))].font=redFont

#This is the final function to be run.  It takes the existing and completed RMS workbook, the number of chips, and the root folder.  It uses RMS
#data to go through and plot ENC vs. peaking time for each channel, and put all those plots in a folder called "ENC Plots" in the root folder.
def ENC_plots(wb,chip_num,plot_dir):
    print ("Plotting ENC data")
    ws = wb.active

#Remove the gain and peaking time, since it's variable
#    gains = ['04.7', '07.8', '14.0', '25.0']
    gains = ['25.0']
    peaking_times = ['0.5', '1.0', '2.0', '3.0']
    plot_peak = [0.5,1,2,3]
    baselines = ['200', '900']
    stuck_text = (.5,-.22)

    try: 
        os.makedirs(plot_dir)
    except OSError:
        if os.path.exists(plot_dir):
            pass

    for chip_id in range(chip_num):
        print ("Chip"+str(chip_id))
        for chn in range(16):
            for index,base in enumerate(baselines):

#Each baseline gets its own plot, since it gets messy if you try to have too many on one
                plt.figure(figsize=(12,8))
                ax = plt.subplot(1,1,1)                          
                ax.set_xlabel("Peaking Time (us)")
                ax.set_ylabel("ENC (electrons)")
                ax.set_title("ENC vs. Peaking Time for Chip "+str(chip_id)+", Channel "+str(chn)+", Baseline = "+str(base)+" mV")
                plot_path = plot_dir+"Chip"+str(chip_id)+"_Channel"+str(chn)+"_Baseline"+str(base)
                for index,gain in enumerate(gains):
                    ENC = []
                    for index,peaking_time in enumerate(peaking_times):

#Build the sheet name from the baseline, gain and peaking time parameter and find it
                        sheet = gain+","+peaking_time+","+base
                        ws = wb[sheet]

#On each sheet, the 6th chunk will have the ENC in electrons, so that channel's value is grabbed, as well as the fill color
                        ENC.append(ws['{}{}'.format(chr(ord('A') + chn + 1),chip_id+4+(6*(chip_num+3)))].value)
                        font_color = ws['{}{}'.format(chr(ord('A') + chn + 1),
                                        chip_id+4+(6*(chip_num+3)))].fill.start_color.index

#If the fill color is charactaristic of a cell that has a stuck bit, plot it with an X.  If not, just a normal scatter plot
                        if (font_color == "00FFC7CE"):
                            ax.scatter(plot_peak[index], ENC[index], marker = 'x', s = 25, color = 'r')
                        else: 
                            ax.scatter(plot_peak[index], ENC[index], marker = '.', s =2, color = 'k')

#Once done, connect the dots
                    ax.plot(plot_peak, ENC, label=str(gain)+" mV/fC Gain")

                # Shrink current axis's height by 10% on the bottom
                box = ax.get_position()
                ax.set_position([box.x0, box.y0 + box.height * 0.1,
                box.width, box.height * 0.9])

#Format the auxilliary information around the plot
                plt.annotate("Red X means there was a stuck bit detected during the RMS analysis", xy=stuck_text,  xycoords='axes fraction',
                xytext=stuck_text, textcoords='axes fraction',
                horizontalalignment='center', verticalalignment='center', color = 'r',
                )
                
                ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.1),ncol=4, fontsize=11)
                plt.savefig (plot_path+".jpg")
                plt.close()
                
def print_pulses(plot_dir, chip_data, value_name):
    for j in range(settings.chip_num):
        time = []
        samples_to_plot = 1000
        for i in range(samples_to_plot):
            time.append(0.5 * i)
        fig = plt.figure(figsize=(16, 12), dpi=80)
        overlay_ax = fig.add_subplot(1,1,1)
        overlay_ax.spines['top'].set_color('none')
        overlay_ax.spines['bottom'].set_color('none')
        overlay_ax.spines['left'].set_color('none')
        overlay_ax.spines['right'].set_color('none')
        overlay_ax.tick_params(labelcolor='w', top='off', bottom='off', left='off', right='off')
        overlay_ax.set_xlabel('Time (us)')
        overlay_ax.set_ylabel('ADC Counts')
        overlay_ax.yaxis.set_label_coords(-0.035,0.5)
        
        ax1 = fig.add_subplot(16,1,16)
        plt.plot(time, chip_data[j][0][0:samples_to_plot])
        plt.setp(ax1.get_xticklabels(), fontsize=12)
        ax1.set_title("Channel 0")
        ax2 = ax1.twinx()
        ax2.set_ylabel("Channel 0", rotation = 0)
        ax2.spines['top'].set_color('none')
        ax2.spines['bottom'].set_color('none')
        ax2.spines['left'].set_color('none')
        ax2.spines['right'].set_color('none')
        ax2.tick_params(labelcolor='w', top='off', bottom='off', left='off', right='off')
    
        for i in range(15):
        
            ax = fig.add_subplot(16,1,15-i, sharex=ax1)
            plt.plot(time, chip_data[j][i+1][0:samples_to_plot])
            plt.setp(ax.get_xticklabels(), visible=False)
            ax2 = ax.twinx()
            ax2.set_ylabel("Channel " + str(i+1), rotation = 0)
    #            ax2.spines['top'].set_color('none')
    #            ax2.spines['bottom'].set_color('none')
    #            ax2.spines['left'].set_color('none')
    #            ax2.spines['right'].set_color('none')
            ax2.tick_params(labelcolor='w', top='off', bottom='off', left='off', right='off')
            pos1 = ax2.get_position() # get the original position 
            pos2 = [pos1.x0 + 0.025, pos1.y0 + 0.005,  pos1.width , pos1.height ] 
            ax2.set_position(pos2) # set a new position
    
        plt.subplots_adjust(wspace=0, hspace=0, top = 1, bottom = 0.05, right = 0.95, left = 0.05)
        directory = plot_dir + "Chip{}".format(j) + "\\"
        try: 
            os.makedirs(directory)
        except OSError:
            if os.path.exists(directory):
                pass
        fig.savefig (directory + "amp_" + value_name) 
        plt.close()
        gc.collect()